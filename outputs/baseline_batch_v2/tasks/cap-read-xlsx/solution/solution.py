import json, os
from openpyxl import load_workbook

IN = '/root/input.xlsx'
OUT = '/root/result.json'

wb = load_workbook(IN, data_only=True)
result = {'sheets': {}, 'sheet_names': []}
for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        row = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                row.append('')
            elif isinstance(v, float):
                if v.is_integer():
                    row.append(str(int(v)))
                else:
                    row.append(str(v))
            else:
                row.append(str(v))
        rows.append(row)
    result['sheets'][name] = rows
    result['sheet_names'].append(name)

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False)
