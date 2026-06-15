import json, os, zipfile

class TestOutputs:
    def test_result_json_exists(self):
        assert os.path.exists('/root/result.json')
        data = json.load(open('/root/result.json', encoding='utf-8'))
        assert data.get('output_path') == '/root/output.xlsx'

    def test_xlsx_exists_and_valid_zip(self):
        p = '/root/output.xlsx'
        assert os.path.exists(p)
        assert zipfile.is_zipfile(p)

    def test_xlsx_has_required_parts(self):
        with zipfile.ZipFile('/root/output.xlsx') as z:
            names = z.namelist()
            assert any(n.endswith('workbook.xml') for n in names)
            assert any('sheet' in n and n.endswith('.xml') for n in names)
            assert any(n.endswith('[Content_Types].xml') or n == '[Content_Types].xml' for n in names)

    def test_structure_fidelity_cells(self):
        from openpyxl import load_workbook
        wb = load_workbook('/root/output.xlsx')
        ws = wb.active
        assert ws.title == 'Sheet1'
        expected = {'A1': '你好', 'B1': '世界', 'A2': '富', 'B2': '酒吧'}
        for coord, val in expected.items():
            assert ws[coord].value == val, (coord, ws[coord].value)

    def test_no_extra_cells(self):
        from openpyxl import load_workbook
        wb = load_workbook('/root/output.xlsx')
        ws = wb.active
        used = [(c.coordinate, c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        assert len(used) == 4
