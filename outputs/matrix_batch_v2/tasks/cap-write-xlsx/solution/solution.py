import json, zipfile, shutil, os
from openpyxl import load_workbook

INPUT = '/root/input.xlsx'
TRANS = '/root/translations.json'
OUT_XLSX = '/root/output.xlsx'
OUT_JSON = '/root/result.json'

translations = json.load(open(TRANS, encoding='utf-8'))

wb = load_workbook(INPUT)
ws = wb.active
# Collect string cells in row-major order matching shared strings index order.
# openpyxl assigns shared string indices on save; simplest: replace cell values directly.
# Build ordered list of string cells.
string_cells = []
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, str):
            string_cells.append(cell)
# Replace by order
for i, cell in enumerate(string_cells):
    if i < len(translations):
        cell.value = translations[i]
wb.save(OUT_XLSX)

# Validate it's a real xlsx
with zipfile.ZipFile(OUT_XLSX) as z:
    names = z.namelist()
    assert any(n.endswith('workbook.xml') for n in names)

with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump({'output_path': OUT_XLSX}, f, ensure_ascii=False)
